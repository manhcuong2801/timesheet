import calendar
import csv
import logging

from openpyxl import load_workbook
from datetime import time, date, datetime, timedelta

month = datetime.now().strftime("%b")

month_num = datetime.now().month
year = datetime.now().year
IN_FILE = (
    f"{year}-0{month_num - 1}-21.csv"
    if month_num < 11
    else f"{year}-{month_num - 1}-21.csv"
)
export_day = IN_FILE[0:10]

SAMPLE_FILE = "[POEMS - Phase 3]- OT PLAN_REQUEST 12_2022.xlsx"
OUT_FILE = "ot-day-poem-3.xlsx"
STATUS_OK = ""
STATUS_DAY_OFF = 1

IGNORE_EMP_ID = [
    "ECO0001",  # CEO
    "ECO0003",  # A. Trung Bom
    "ECO0012",  # Ma Nhu
    "ECO0038",  # Chi Chung
    "ECO0089",  # Vuong
    "ECO0345",  # Co Sam
]


def compute_actual_working_hours(
    checkin_time: datetime, work_hours: float, time_late: float
):
    real_work_hours = 0
    day_of_week_str = f"{str(checkin_time)[0:10]}"
    end_time_str = f"{str(checkin_time)[0:10]} 17:30:00"
    after_work_time_str = f"{str(checkin_time)[0:10]} 18:30:00"
    check_out_str = f"{str(check_out)}"
    day_of_week = datetime.strptime(day_of_week_str, "%Y-%m-%d")
    time_checkout = datetime.strptime(check_out_str, datetimeFormat)
    end_time = str2datetime(end_time_str)
    after_work_time = str2datetime(after_work_time_str)
    if day_of_week.weekday() <= 4:
        real_work_hours = float(work_hours) - 9 - float(time_late)
    else:
        if time_checkout < end_time:
            real_work_hours = float(work_hours) - 8 - float(time_late)
        elif time_checkout.hour - after_work_time.hour > 0:
            real_work_hours = float(work_hours) - float(time_late)

    return real_work_hours if real_work_hours > 0 else ""


def time2str(value):
    if isinstance(value, str):
        return value
    if isinstance(value, time):
        return value.strftime("%d")
    return ""


def str2date(value):
    if isinstance(value, (date, datetime)):
        return value
    return datetime.strptime(value, "%Y-%m-%d")


def str2datetime(value):
    if isinstance(value, (date, datetime)):
        return value
    return datetime.strptime(value, "%Y-%m-%d %H:%M:%S")


def get_late_time(time: float) -> float:
    late_time: float
    if time >= 8:
        return 8
    if time <= 0.25:
        late_time = 0.25
    elif 0.25 < time <= 0.5:
        late_time = 0.5
    elif 0.5 < time <= 0.75:
        late_time = 0.75
    else:
        late_time = time
    return late_time


datetimeFormat = "%Y-%m-%d %H:%M:%S"
first_day = time2str(f"{export_day}")
start = datetime.strptime(first_day, "%Y-%m-%d")

days_in_month = calendar.monthrange(start.year, start.month)[1]
plus_day = timedelta(days=days_in_month - 1)
end = start + plus_day
step = timedelta(days=1)

data = {}

working_days = []


print(f"Start loading template from {SAMPLE_FILE}")
out_wb = load_workbook(SAMPLE_FILE)
out_ws = out_wb.worksheets[3]

k = 0  # column
row = 3

data_col = set()
while row < 290:
    a = out_ws.cell(row=row, column=2).value
    if a:
        data_col.add(a)
    row += 1


while start <= end:
    column = 7 + k
    out_ws.cell(row=1, column=column, value=start.strftime("%d/%m/%Y"))
    working_days.append(start)
    k += 10
    start += step
print(working_days)

with open(IN_FILE, mode="r", encoding="utf-8") as file:
    reader = csv.reader(file)
    try:
        for row in reader:
            emp_id = row[0].upper()
            emp_name = row[1].upper()
            key = (emp_id, emp_name)
            if emp_id not in data_col:
                continue
            if key not in data:
                data[key] = []
            day = str2datetime(row[2])
            check_in = row[2]
            check_out = row[3]
            late_time = row[4]
            working_hours = row[5]
            data[key].insert(0, (day, check_in, check_out, late_time, working_hours))
            # print(data[key])
    except Exception as e:
        logging.exception('oi doi oi')
if not data:
    print('No data')

time_now = f"{month}_{datetime.now().year}"

step_row = 1
status_checkin = ""
status_checkout = ""
status_late = ""
status_work = ""
status_actual = ""


for emp_id, emp_name in sorted(data):
    zz = 0
    check_time = data[(emp_id, emp_name)]
    if not check_time:
        continue
    row_id = 3
    start_day = datetime.strptime(first_day, "%Y-%m-%d")
    for day, check_in, check_out, late_time, working_hours in check_time:
        column = 8 + zz
        id_emp = out_ws.cell(row=row_id, column=2).value
        column_date_str = out_ws.cell(row=1, column=column - 1).value
        column_date = datetime.strptime(column_date_str, "%d/%m/%Y")
        while id_emp != emp_id:
            row_id += step_row
            id_emp = out_ws.cell(row=row_id, column=2).value

        while column_date < datetime.strptime(str(day)[0:10], "%Y-%m-%d"):
            column_date += step
            column += 10
        if column_date.weekday() > 5:
            continue
        elif day.strftime("%d") != column_date.strftime("%d"):
            continue
        elif day.strftime("%d/%m/%Y") == column_date.strftime("%d/%m/%Y"):
            status_checkin = check_in
            status_checkout = check_out
            status_late = late_time
            status_work = working_hours
            # status_actual = compute_actual_working_hours(
            #     check_in, working_hours, late_time
            # )

        cell_checkin = out_ws.cell(row=row_id, column=column, value=status_checkin)
        cell_checkout = out_ws.cell(
            row=row_id, column=column + 1, value=status_checkout
        )
        cell_late = out_ws.cell(row=row_id, column=column + 2, value=status_late)
        cell_work = out_ws.cell(row=row_id, column=column + 3, value=status_work)
        # cell_actual = out_ws.cell(row=row_id, column=column + 4)
        zz += 10
        start_day += step
        print(f"{day},  {emp_name}, {row_id}")
    row_id += step_row


print(f"Saving working day to {OUT_FILE}")
out_wb.save(OUT_FILE)
print("DONE.")
