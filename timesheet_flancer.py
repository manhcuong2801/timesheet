import calendar
import csv

from openpyxl import load_workbook
from openpyxl.comments import Comment
from datetime import time, date, datetime, timedelta
import datetime as dt


NO_CHECKOUT = "NO CHECKOUT"

month = datetime.now().strftime("%b")

month_num = datetime.now().month
year = datetime.now().year
START_DAY = (
    f"{year}-0{month_num - 1}-21.csv"
    if month_num < 10
    else f"{year}-{month_num-1}-21.csv"
)
export_day = START_DAY[0:10]
IN_FILE = "free_lancer.csv"
# export_day = '2021-12-21'
SAMPLE_FILE = "WorkingdayTemplate.xlsx"
OUT_FILE = "Workingday.xlsx"

STATUS_OK = ""
STATUS_DAY_OFF = 1

IGNORE_EMP_ID = [
    "ECO0001",  # CEO
    "ECO0003",  # A. Trung Bom
    "ECO0012",  # Ma Nhu
    "ECO0038",  # Chi Chung
    "ECO0089",  # Vuong
    "ECO0345",  # Co Sam
]


def time2str(value):
    if isinstance(value, str):
        return value
    if isinstance(value, time):
        return value.strftime("%d")
    return ""


def str2date(value):
    if isinstance(value, (date, datetime)):
        return value
    return datetime.strptime(value, "%Y-%m-%d")


def str2datetime(value):
    if isinstance(value, (date, datetime)):
        return value
    return datetime.strptime(value, "%Y-%m-%d %H:%M:%S")


def get_late_time(value: float, digits: int = 2) -> float:

    if value >= 8:
        return 8
    number_digits = 10 ** digits
    return round(value * number_digits) / number_digits


datetimeFormat = "%Y-%m-%d %H:%M:%S"
first_day = time2str(f"{export_day}")
start = datetime.strptime(first_day, "%Y-%m-%d")

days_in_month = calendar.monthrange(start.year, start.month)[1]
plus_day = timedelta(days=days_in_month - 1)
end = start + plus_day
step = timedelta(days=1)

data = {}

working_days = []


print(f"Start loading template from {SAMPLE_FILE}")
out_wb = load_workbook(SAMPLE_FILE)
out_ws = out_wb.active

k = 0  # column
while start <= end:
    column = 6 + k
    out_ws.cell(row=6, column=column, value=start.strftime("%a"))
    out_ws.cell(row=7, column=column, value=start.day)
    working_days.append(start)
    k += 1
    start += step
# print(working_days)

with open(IN_FILE, mode="r", encoding="utf-8") as file:
    reader = csv.reader(file)
    for row in reader:
        emp_id = row[0].upper()
        emp_name = row[1].upper()
        key = (emp_id, emp_name)
        if key not in data:
            data[key] = []
        day = str2datetime(row[2])
        checkout_time = str2datetime(row[3]) if row[3] else NO_CHECKOUT
        right_time_str = f"{str(row[2])[0:10]} 08:30:00"
        noon_time_str = f"{str(row[2])[0:10]} 13:00:00"
        right_co_time_str = f"{str(row[2])[0:10]} 17:30:00"
        right_co_time = str2datetime(right_co_time_str)
        late_time_raw = 0
        late_time_real = 0

        if str2datetime(row[2]) > str2datetime(right_time_str):
            late_time_raw = str2datetime(row[2]) - str2datetime(right_time_str)

        if checkout_time < str2datetime(right_co_time_str):
            late_time_real += (right_co_time - checkout_time).seconds / 3600
        if late_time_raw:
            late_time_real = late_time_raw.seconds / 3600
            if late_time_raw.seconds > 16200:
                late_time_raw = (
                    str2datetime(row[2])
                    - str2datetime(right_time_str)
                    + timedelta(hours=4)
                ).seconds
        late_time = get_late_time(late_time_real)
        data[key].append((day, late_time, late_time_real, checkout_time))
        print(data[key])


print(f"Start loading template from {SAMPLE_FILE}")
out_wb = load_workbook(SAMPLE_FILE)
out_ws = out_wb.active
k = 0  # column
start_day = datetime.strptime(first_day, "%Y-%m-%d")
time_now = f"{month}_{datetime.now().year}"
out_ws.cell(row=2, column=1).value = f"TIME ATTENDANCE RECORD {time_now}"
from_day = export_day.replace("-", "/")
end_day = str(end)[0:10].replace("-", "/")
out_ws.cell(row=4, column=1).value = f"WORKING DAY {from_day} - {end_day}"
while start_day <= end:
    column = 6 + k
    out_ws.cell(row=6, column=column, value=start_day.strftime("%a"))
    out_ws.cell(row=7, column=column, value=start_day.day)

    i = 0  # row
    for emp_id, emp_name in sorted(data):

        status = STATUS_OK
        row = 8 + i
        out_ws.cell(row=row, column=1).value = i + 1
        out_ws.cell(row=row, column=2).value = emp_id
        out_ws.cell(row=row, column=3).value = emp_name

        check_time = data[(emp_id, emp_name)]
        for day, late_time, late_time_real, checkout_time in check_time:
            # Column start from 6 because there 2 hidden columns
            comments = []
            comment_tool = ""
            cell = out_ws.cell(row=row, column=column)
            time = day.strftime("%H:%M:%S")

            comments.append(f"Checkout lúc: {checkout_time}")
            if emp_id in IGNORE_EMP_ID:
                status = STATUS_OK
            elif start_day.weekday() > 5:
                status = STATUS_OK
                comments = []
            elif day.strftime("%d") != start_day.strftime("%d"):
                comments = []
                continue
            elif late_time_real > 0 and late_time > 0:
                status = late_time * 0.125
                comments.append(f"Checkin lúc: {time}")
            else:
                status = 0
            if status != "":
                cell.value = status
                print(f"---{emp_name}---{cell} --- {status}")
            if comments:
                comments.insert(0, f"TDT STAFF: ")
                comment_tool = "\n".join(comments)
                cell.comment = Comment(comment_tool, "Tool")
                comments = []
        sum_cell = out_ws.cell(row=row, column=38)
        sum_cell.value = f"=SUM(F{row}:AJ{row})"

        i += 1

    k += 1
    start_day += step

print(f"Saving working day to {OUT_FILE}")
out_wb.save(OUT_FILE)
print("DONE.")
