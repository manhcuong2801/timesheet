import calendar
import glob
import os

from openpyxl import load_workbook
from datetime import datetime, timedelta

from OT import utils
from OT.utils import time2str, STEP

month = datetime.now().strftime("%b")
ROOT_PATH = '/home/cuonglm/PSP/new-timesheet/OT/sample_files/'
RESULT_PATH = '/home/cuonglm/PSP/new-timesheet/OT/result/'
month_num = datetime.now().month
year = datetime.now().year
IN_FILE = (
    f"{year}-0{month_num - 1}-21.csv"
    if month_num < 11
    else f"{year}-{month_num - 1}-21.csv"
)
export_day = IN_FILE[0:10]

SAMPLE_FILE = "AFX-OT-PLAN_REQUEST-12_2022.xlsx"
OUT_FILE = "ot-day.xlsx"
STATUS_OK = ""
STATUS_DAY_OFF = 1

IGNORE_EMP_ID = [
    "ECO0001",  # CEO
    "ECO0003",  # A. Trung Bom
    "ECO0012",  # Ma Nhu
    "ECO0038",  # Chi Chung
    "ECO0089",  # Vuong
    "ECO0345",  # Co Sam
]


datetimeFormat = "%Y-%m-%d %H:%M:%S"
first_day = time2str(f"{export_day}")


def generate_ot_sheet(sample_file):
    start = datetime.strptime(first_day, "%Y-%m-%d")
    days_in_month = calendar.monthrange(start.year, start.month)[1]
    plus_day = timedelta(days=days_in_month - 1)
    end = start + plus_day
    k = 0  # column
    working_days = []
    try:
        out_wb = load_workbook(sample_file)
        sheets_list = out_wb.sheetnames
        idx_work_sheet = sheets_list.index('2111-2012')
        out_ws = out_wb.worksheets[idx_work_sheet]

        while start <= end:
            column = 7 + k
            out_ws.cell(row=1, column=column, value=start.strftime("%d/%m/%Y"))
            working_days.append(start)
            k += 10
            start += STEP

        data_col = utils.handle_getting_from_sample_file(out_ws=out_ws)
        data = utils.read_from_file(in_file=IN_FILE, data_col=data_col)

        utils.handle_fill_data(data=data, first_day=first_day, out_ws=out_ws)

        # get basename
        out_file = os.path.basename(sample_file)
        path_result = f"{RESULT_PATH}/{out_file}"
        print(f"Saving working day to {path_result}")
        out_wb.save(path_result)
        print("DONE.")
    except Exception as e:
        print(f'thoi bo me roi: {sample_file}')


all_sample_file = [f for f in glob.glob("/home/cuonglm/PSP/new-timesheet/OT/sample_files/*")]

for sample in all_sample_file:
    print("START")
    generate_ot_sheet(sample)
print("DONE ALL")
